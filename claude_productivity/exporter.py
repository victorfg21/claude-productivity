"""
Exportação para planilha Excel (.xlsx) formatada para uso gerencial.

Abas geradas:
  1. Resumo       — visão geral da sessão e métricas chave
  2. Histórico    — atividade diária dos últimos 7 dias
  3. Projetos     — métricas agregadas por projeto
  4. Linguagens   — breakdown de linguagens por projeto
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

from .db import DailyStats, ProjectStats, SessionStats

# ── Paleta de cores (hexadecimais sem #) ────────────────────────────────────

_C = {
    "header_bg":   "1C2128",   # fundo cabeçalho
    "header_fg":   "61AFEF",   # texto cabeçalho azul
    "title_bg":    "21252B",   # fundo título de seção
    "title_fg":    "98C379",   # verde
    "row_alt":     "282C34",   # linha alternada escura
    "row_base":    "1D2026",   # linha base
    "accent":      "61AFEF",   # azul principal
    "warning":     "E06C75",   # vermelho
    "success":     "98C379",   # verde
    "tip":         "E5C07B",   # amarelo
    "text":        "ABB2BF",   # texto normal
    "dim":         "5C6370",   # texto apagado
    "white":       "FFFFFF",
}


def _font(bold=False, color="ABB2BF", size=11) -> Font:
    return Font(name="Consolas", bold=bold, color=color, size=size)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border_bottom(color="3E4451") -> Border:
    side = Side(style="thin", color=color)
    return Border(bottom=side)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _write_header_row(ws, row: int, cols: list[str], widths: list[int]) -> None:
    """Escreve linha de cabeçalho de tabela com estilo."""
    ws.row_dimensions[row].height = 22
    for ci, (label, width) in enumerate(zip(cols, widths), start=1):
        cell = ws.cell(row=row, column=ci, value=f"  {label}" if ci == 1 else label)
        cell.font      = _font(bold=True, color=_C["header_fg"], size=10)
        cell.fill      = _fill(_C["header_bg"])
        cell.alignment = _align("center" if ci > 1 else "left", "center")
        cell.border    = _border_bottom(_C["accent"])
        ws.column_dimensions[get_column_letter(ci)].width = width


def _write_section_title(ws, row: int, title: str, span: int) -> None:
    """Escreve um título de seção em destaque."""
    cell = ws.cell(row=row, column=1, value=f"  {title}")
    cell.font      = _font(bold=True, color=_C["title_fg"], size=11)
    cell.fill      = _fill(_C["title_bg"])
    cell.alignment = _align("left")
    if span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def _data_cell(ws, row: int, col: int, value, bold=False, color=None,
               align="left", number_format=None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _font(bold=bold, color=color or _C["text"])
    cell.alignment = _align(align)
    bg = _C["row_alt"] if row % 2 == 0 else _C["row_base"]
    cell.fill      = _fill(bg)
    if number_format:
        cell.number_format = number_format


# ── Aba 1 — Resumo ───────────────────────────────────────────────────────────

def _build_resumo(ws, session: SessionStats, now: datetime) -> None:
    ws.sheet_view.showGridLines = False

    # Título principal
    ws.row_dimensions[1].height = 34
    title = ws.cell(row=1, column=1, value="Claude Code — Relatório de Produtividade")
    title.font      = _font(bold=True, color=_C["white"], size=14)
    title.fill      = _fill(_C["header_bg"])
    title.alignment = _align("left", "center")
    ws.merge_cells("A1:D1")

    subtitle = ws.cell(row=1, column=5,
                       value=f"Gerado em {now.strftime('%d/%m/%Y %H:%M')}")
    subtitle.font      = _font(color=_C["dim"], size=10)
    subtitle.fill      = _fill(_C["header_bg"])
    subtitle.alignment = _align("right", "center")
    ws.merge_cells("E1:G1")

    ws.row_dimensions[2].height = 8

    # Seção: Sessão Atual
    _write_section_title(ws, 3, "SESSÃO ATUAL", 7)
    ws.row_dimensions[3].height = 22

    RESUMO_ROWS = [
        ("Projeto",          session.project_name or "—",        False, _C["accent"]),
        ("Status",           "Ativa" if session.is_active else "Encerrada",
                             True,
                             _C["success"] if session.is_active else _C["dim"]),
        ("Duração",          f"{int(session.duration_minutes // 60)}h {int(session.duration_minutes % 60)}m",
                             False, _C["text"]),
        ("Total de ações",   session.total_tools,                True, _C["accent"]),
        ("Arquivos únicos",  session.unique_files,               False, _C["text"]),
        ("",                 "",                                 False, _C["text"]),
        ("Edits / Writes",   session.edit_count,                 False, _C["tip"]),
        ("Comandos Bash",    session.bash_count,                 False, _C["text"]),
        ("Leituras",         session.read_count,                 False, _C["text"]),
        ("Subagents usados", session.agent_calls,                False, _C["text"]),
        ("",                 "",                                 False, _C["text"]),
        ("Bash — Taxa de sucesso",
         f"{session.bash_success_rate:.1f}%" if session.bash_success_rate > 0 else "—",
         True,
         _C["success"] if session.bash_success_rate >= 80
         else (_C["warning"] if 0 < session.bash_success_rate < 70 else _C["tip"])),
        ("Burst médio de edits",
         f"{session.avg_edit_burst:.1f} edits consecutivos" if session.avg_edit_burst > 0 else "—",
         False, _C["text"]),
        ("Arquivos retomados (cross-session)",
         len(session.cross_session_files), False, _C["text"]),
    ]

    for i, (label, value, bold, color) in enumerate(RESUMO_ROWS, start=4):
        r = i
        ws.row_dimensions[r].height = 20
        lbl_cell = ws.cell(row=r, column=1, value=f"  {label}" if label else "")
        lbl_cell.font      = _font(color=_C["dim"], size=10)
        lbl_cell.fill      = _fill(_C["row_alt"] if r % 2 == 0 else _C["row_base"])
        lbl_cell.alignment = _align("left", "center")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

        val_cell = ws.cell(row=r, column=4, value=value)
        val_cell.font      = _font(bold=bold, color=color, size=11)
        val_cell.fill      = _fill(_C["row_alt"] if r % 2 == 0 else _C["row_base"])
        val_cell.alignment = _align("left", "center")
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)

    # Seção: Linguagens
    r = 4 + len(RESUMO_ROWS) + 1
    _write_section_title(ws, r, "LINGUAGENS DA SESSÃO", 7)
    ws.row_dimensions[r].height = 22
    r += 1

    if session.language_breakdown:
        total_ext = sum(session.language_breakdown.values())
        _write_header_row(ws, r, ["Extensão", "Arquivos editados", "% do total"], [18, 24, 16])
        r += 1
        for ext, cnt in session.language_breakdown.items():
            pct = cnt / total_ext * 100
            ws.row_dimensions[r].height = 20
            _data_cell(ws, r, 1, f"  .{ext}", bold=True, color=_C["accent"])
            _data_cell(ws, r, 2, cnt, align="center")
            _data_cell(ws, r, 3, f"{pct:.1f}%", align="center",
                       color=_C["success"] if pct >= 50 else _C["text"])
            r += 1
    else:
        ws.row_dimensions[r].height = 20
        ws.cell(row=r, column=1, value="  Nenhum dado de linguagem registrado").font = _font(color=_C["dim"])

    # Colunas
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18

    # Congela linha de cabeçalho
    ws.freeze_panes = "A4"


# ── Aba 2 — Histórico ────────────────────────────────────────────────────────

def _build_historico(ws, history: List[DailyStats]) -> None:
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 34
    title = ws.cell(row=1, column=1, value="Histórico de Atividade — Últimos 7 Dias")
    title.font      = _font(bold=True, color=_C["white"], size=14)
    title.fill      = _fill(_C["header_bg"])
    title.alignment = _align("left", "center")
    ws.merge_cells("A1:H1")

    ws.row_dimensions[2].height = 8

    COLS  = ["Data", "Dia da Semana", "Total de Ações", "Edits / Writes",
             "Comandos Bash", "Leituras", "Arquivos Únicos", "Tempo Ativo"]
    WIDTHS = [18, 20, 22, 22, 20, 16, 20, 18]
    _write_header_row(ws, 3, COLS, WIDTHS)
    ws.freeze_panes = "A4"

    today = datetime.now().strftime("%Y-%m-%d")
    DAYS_PT = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]

    for i, d in enumerate(history, start=4):
        ws.row_dimensions[i].height = 20
        is_today  = d.date == today
        is_active = d.total_tools > 0
        row_color = _C["title_bg"] if is_today else (_C["row_alt"] if i % 2 == 0 else _C["row_base"])

        try:
            dt   = datetime.strptime(d.date, "%Y-%m-%d")
            dow  = DAYS_PT[dt.weekday()]
        except ValueError:
            dow  = "—"

        mins = int(d.active_minutes)
        time_str = f"{mins // 60}h {mins % 60:02d}m" if mins > 0 else "—"

        cells_data = [
            (d.date,          True if is_today else False,  _C["accent"] if is_today else _C["text"]),
            (dow,             False,                         _C["dim"]),
            (d.total_tools,   True,                          _C["accent"] if is_active else _C["dim"]),
            (d.edit_count,    False,                         _C["tip"] if is_active else _C["dim"]),
            (d.bash_count,    False,                         _C["text"] if is_active else _C["dim"]),
            (d.read_count if hasattr(d, "read_count") else "—",
             False, _C["dim"]),
            (d.unique_files,  False,                         _C["text"] if is_active else _C["dim"]),
            (time_str,        False,                         _C["text"] if is_active else _C["dim"]),
        ]
        for col_i, (val, bold, color) in enumerate(cells_data, start=1):
            cell = ws.cell(row=i, column=col_i, value=f"  {val}" if col_i == 1 else val)
            cell.font      = _font(bold=bold, color=color)
            cell.fill      = _fill(row_color)
            cell.alignment = _align("center" if col_i > 1 else "left", "center")

    # Totais / médias
    active_days = [d for d in history if d.total_tools > 0]
    if active_days:
        r = 4 + len(history) + 1
        ws.row_dimensions[r].height = 6
        r += 1
        _write_section_title(ws, r, "MÉDIAS (dias com atividade)", 8)
        r += 1

        avg_tools = sum(d.total_tools for d in active_days) / len(active_days)
        avg_edits = sum(d.edit_count  for d in active_days) / len(active_days)
        avg_bash  = sum(d.bash_count  for d in active_days) / len(active_days)
        avg_mins  = sum(d.active_minutes for d in active_days) / len(active_days)
        avg_time  = f"{int(avg_mins) // 60}h {int(avg_mins) % 60:02d}m"

        summary = [
            ("Média de ações/dia",  f"{avg_tools:.0f}",  _C["accent"]),
            ("Média de edits/dia",  f"{avg_edits:.0f}",  _C["tip"]),
            ("Média de bash/dia",   f"{avg_bash:.0f}",   _C["text"]),
            ("Tempo médio/dia",     avg_time,             _C["text"]),
            ("Dias ativos",         f"{len(active_days)}/{len(history)}", _C["success"]),
        ]
        for label, val, color in summary:
            ws.row_dimensions[r].height = 20
            lbl = ws.cell(row=r, column=1, value=f"  {label}")
            lbl.font = _font(color=_C["dim"])
            lbl.fill = _fill(_C["row_base"])
            lbl.alignment = _align("left", "center")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

            v = ws.cell(row=r, column=4, value=val)
            v.font = _font(bold=True, color=color)
            v.fill = _fill(_C["row_base"])
            v.alignment = _align("left", "center")
            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
            r += 1


# ── Aba 3 — Projetos ─────────────────────────────────────────────────────────

def _build_projetos(ws, projects: List[ProjectStats]) -> None:
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 34
    title = ws.cell(row=1, column=1, value="Métricas por Projeto")
    title.font      = _font(bold=True, color=_C["white"], size=14)
    title.fill      = _fill(_C["header_bg"])
    title.alignment = _align("left", "center")
    ws.merge_cells("A1:J1")

    ws.row_dimensions[2].height = 8

    COLS   = ["Projeto", "Último uso", "Sessões", "Edits / Writes",
              "Comandos Bash", "Arquivos Únicos", "Tempo Total",
              "Bash Sucesso", "Linguagens Principais", "Leituras"]
    WIDTHS = [30, 18, 12, 20, 20, 20, 18, 16, 34, 14]
    _write_header_row(ws, 3, COLS, WIDTHS)
    ws.freeze_panes = "A4"

    for i, pr in enumerate(projects, start=4):
        ws.row_dimensions[i].height = 22
        mins     = int(pr.total_minutes)
        time_str = f"{mins // 60}h {mins % 60:02d}m" if mins > 0 else "—"
        bash_str = f"{pr.bash_success_rate:.0f}%" if pr.bash_success_rate > 0 else "—"
        bash_color = (
            _C["success"] if pr.bash_success_rate >= 80
            else (_C["warning"] if 0 < pr.bash_success_rate < 70 else _C["tip"])
        ) if pr.bash_success_rate > 0 else _C["dim"]
        lang_str = "  /  ".join(
            f".{ext} ({cnt})" for ext, cnt in list(pr.language_breakdown.items())[:3]
        ) or "—"

        cells = [
            (pr.project_name,    True,  _C["accent"]),
            (pr.last_seen,       False, _C["dim"]),
            (pr.total_sessions,  False, _C["text"]),
            (pr.total_edits,     True,  _C["tip"]),
            (pr.total_bash,      False, _C["text"]),
            (pr.unique_files,    False, _C["text"]),
            (time_str,           False, _C["text"]),
            (bash_str,           True,  bash_color),
            (lang_str,           False, _C["dim"]),
            (pr.total_reads,     False, _C["dim"]),
        ]
        row_bg = _C["row_alt"] if i % 2 == 0 else _C["row_base"]
        for ci, (val, bold, color) in enumerate(cells, start=1):
            display_val = f"  {val}" if ci == 1 else val
            cell = ws.cell(row=i, column=ci, value=display_val)
            cell.font      = _font(bold=bold, color=color)
            cell.fill      = _fill(row_bg)
            cell.alignment = _align("center" if ci > 1 else "left", "center", wrap=(ci == 9))


# ── Aba 4 — Linguagens ───────────────────────────────────────────────────────

def _build_linguagens(ws, projects: List[ProjectStats]) -> None:
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 34
    title = ws.cell(row=1, column=1, value="Breakdown de Linguagens por Projeto")
    title.font      = _font(bold=True, color=_C["white"], size=14)
    title.fill      = _fill(_C["header_bg"])
    title.alignment = _align("left", "center")
    ws.merge_cells("A1:E1")

    ws.row_dimensions[2].height = 8

    COLS   = ["Projeto", "Extensão", "Arquivos editados", "% no projeto", "Nível de uso"]
    WIDTHS = [32, 16, 24, 18, 28]
    _write_header_row(ws, 3, COLS, WIDTHS)
    ws.freeze_panes = "A4"

    r = 4
    for pr in projects:
        if not pr.language_breakdown:
            continue
        total = sum(pr.language_breakdown.values())
        for ext, cnt in sorted(pr.language_breakdown.items(), key=lambda x: -x[1]):
            pct  = cnt / total * 100 if total else 0
            bar  = "█" * int(pct / 5) + "░" * (20 - int(pct / 5))
            row_bg = _C["row_alt"] if r % 2 == 0 else _C["row_base"]
            ws.row_dimensions[r].height = 20

            cells = [
                (pr.project_name, False, _C["dim"]),
                (f".{ext}",       True,  _C["accent"]),
                (cnt,             False, _C["text"]),
                (f"{pct:.1f}%",   True,  _C["success"] if pct >= 50 else _C["text"]),
                (bar,             False, _C["tip"]),
            ]
            for ci, (val, bold, color) in enumerate(cells, start=1):
                display_val = f"  {val}" if ci == 1 else val
                cell = ws.cell(row=r, column=ci, value=display_val)
                cell.font      = _font(bold=bold, color=color, size=10 if ci == 5 else 11)
                cell.fill      = _fill(row_bg)
                cell.alignment = _align("center" if ci > 1 else "left", "center")
            r += 1
        # Linha de separação entre projetos
        ws.row_dimensions[r].height = 4
        r += 1


# ── Entry point ──────────────────────────────────────────────────────────────

def export_xlsx(
    session: SessionStats,
    history: List[DailyStats],
    projects: List[ProjectStats],
) -> Path:
    """
    Gera a planilha Excel e salva em ~/claude-metrics-YYYYMMDD-HHMMSS.xlsx.
    Retorna o caminho do arquivo criado.
    """
    now = datetime.now()
    wb  = Workbook()

    # Remove a aba padrão
    wb.remove(wb.active)

    # Cria as abas
    ws_resumo    = wb.create_sheet("Resumo")
    ws_historico = wb.create_sheet("Histórico 7 Dias")
    ws_projetos  = wb.create_sheet("Projetos")
    ws_lang      = wb.create_sheet("Linguagens")

    # Define cor de aba (tab color)
    ws_resumo.sheet_properties.tabColor    = "61AFEF"
    ws_historico.sheet_properties.tabColor = "98C379"
    ws_projetos.sheet_properties.tabColor  = "E5C07B"
    ws_lang.sheet_properties.tabColor      = "C678DD"

    # Preenche
    _build_resumo(ws_resumo, session, now)
    _build_historico(ws_historico, history)
    _build_projetos(ws_projetos, projects)
    _build_linguagens(ws_lang, projects)

    # Salva
    out_path = Path.home() / f"claude-metrics-{now.strftime('%Y%m%d-%H%M%S')}.xlsx"
    wb.save(out_path)
    return out_path
